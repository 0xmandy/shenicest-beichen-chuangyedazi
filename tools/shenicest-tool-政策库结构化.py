# -*- coding: utf-8 -*-
"""由 Claude 生成，属于 shenicest 黑客松北辰命题原型。
把《政策智能体数据库.xlsx》两张表结构化成前端可用的 JS 数据块。
字段只做两件事：原样保留（名称/日期/URL/接入状态），以及从标题正则派生（层级/主题标签/事项类型）。
不派生额度、截止日、申报条件 —— 原表没有，编出来就是幻觉。"""
import json, re, os

import openpyxl

# 源表不进 repo（是命题给的原始材料）。跑之前把 xlsx 放到 data/ 下，或者用 POLICY_XLSX 指过来。
_HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.environ.get('POLICY_XLSX') or os.path.join(_HERE, '..', 'data', '政策智能体数据库.xlsx')
OUT = os.environ.get('POLICY_OUT') or os.path.join(_HERE, 'policy-db.js')
if not os.path.exists(SRC):
    raise SystemExit('找不到政策库源表：%s\n把 xlsx 放到 data/ 下，或者 POLICY_XLSX=/path/to.xlsx 再跑' % SRC)

TAG_RULES = [
    ('人工智能', ['人工智能', 'AI', '大模型', '智能体', '通用人工智能', 'OPC']),
    ('智能机器人', ['机器人', '具身智能', '人形']),
    ('数据要素', ['数据要素', '公共数据', '数据知识产权', '数据跨境', '数据出境', '数据资源']),
    ('知识产权', ['知识产权', '专利', '商标', '著作权']),
    ('融资金融', ['金融', '融资', '贷款', '贴息', '风险补偿', '质押', '普惠']),
    ('中小企业', ['中小企业', '专精特新', '梯度培育', '小微', '中小微']),
    ('人才支持', ['人才', '科技新星', '凤凰计划', '留学回国', '技能大师', '职称', '青年科技']),
    ('场景开放', ['场景', '揭榜', '示范项目', '实训']),
    ('互联网3.0', ['互联网 3.0', '互联网3.0', 'XR', '数字空间', '区块链', '超高清', '视听']),
    ('医药健康', ['医疗', '医药', '健康', '器械', '药品', '康养']),
    ('园区载体', ['孵化器', '服务载体', '服务站', '园区', '科技百园', '开发区', '商务中心区']),
    ('成果转化', ['成果转化', '技术转移', '概念验证', '重点实验室', '创新券', '技术市场', '中试']),
    ('企业出海', ['出海', '跨境', '自由贸易', '自贸']),
    ('信息软件', ['信息软件', '软件', '集成电路', '工业互联网', '互联网基础资源', '6G']),
    ('商务消费', ['商务', '总部企业', '消费', '以旧换新', '预付卡', '文旅', '游戏', '电竞']),
    ('高新技术', ['高新技术', '高精尖', '创新能力', '科技计划', '研发费用', '科学技术奖']),
    ('绿色制造', ['绿色', '节能', '降碳', '智能工厂', '数字化转型', '新型工业化']),
]

LEVEL_RULES = [
    ('朝阳区', ['朝阳区', '朝阳向新', '朝阳十五条']),
    ('中关村', ['中关村']),
    ('经开区', ['北京经济技术开发区', '经济技术开发区']),
    ('门头沟区', ['门头沟']),
    ('通州区', ['通州', '城市副中心']),
    ('昌平区', ['昌平']),
    ('北京市', ['北京市', '北京', '首都', '本市', '京市']),
]

TYPE_RULES = [
    ('申报通知', ['申报', '征集', '揭榜', '奖励申报', '评审', '公告']),
    ('实施细则', ['实施细则', '实施方案', '实施办法', '工作指引', '指南']),
    ('管理办法', ['管理办法', '办法', '规定', '条例', '标准']),
    ('行动计划', ['行动计划', '行动方案', '工作方案', '规划']),
    ('支持措施', ['若干措施', '措施', '支持政策', '意见']),
]

# 能直接对上「企业要办的事」的强信号：这类给可申报标记
ACTIONABLE = ['申报', '征集', '揭榜', '奖励', '认定', '资助', '补贴', '贴息', '资金', '支持资金']


def pick_tags(name):
    tags = []
    for tag, kws in TAG_RULES:
        if any(k in name for k in kws):
            tags.append(tag)
    return tags[:3] or ['综合政策']


def pick_level(name):
    for lv, kws in LEVEL_RULES:
        if any(k in name for k in kws):
            return lv
    return '其他'


def pick_type(name):
    for tp, kws in TYPE_RULES:
        if any(k in name for k in kws):
            return tp
    return '政策文件'


def clean_name(raw):
    n = re.sub(r'\.(pdf|docx?|xlsx?|doc|xls)$', '', str(raw).strip(), flags=re.I)
    n = n.replace('"', '「', 1).replace('"', '」', 1)
    return n.strip()


def norm(s):
    return re.sub(r'[《》（）()\s"“”\-—·、,，。\.]+', '', s)


wb = openpyxl.load_workbook(SRC, data_only=True)

# ---- 库一：全文导入的政策原文 ----
docs = []
ws = wb['整个文本导入接入']
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[1] or str(r[1]).strip() in ('Q&A',):
        continue
    name = clean_name(r[1])
    docs.append({
        'id': 'P%03d' % (len(docs) + 1),
        'name': name,
        'file': str(r[1]).strip(),
        'level': pick_level(name),
        'tags': pick_tags(name),
        'type': pick_type(name),
        'link': str(r[2] or '导入').strip(),
        'status': str(r[3] or '').strip(),
        'updated': str(r[4] or '')[:10],
        'reads': [],
    })

# ---- 库二：URL 接入的政策与解读 ----
reads = []
ws2 = wb['只有url接入到智能体']
for r in ws2.iter_rows(min_row=2, values_only=True):
    if not r[1] or not r[3]:
        continue
    title = str(r[1]).strip()
    reads.append({
        'id': 'R%03d' % (len(reads) + 1),
        'title': title,
        'date': str(r[2] or '')[:10],
        'url': str(r[3]).strip(),
        'level': pick_level(title),
        'tags': pick_tags(title),
        'type': pick_type(title),
        'actionable': any(k in title for k in ACTIONABLE),
    })

# ---- 把解读挂回原文（标题归一化后互为子串即认定同一份）----
hit = 0
for rd in reads:
    inner = re.findall(r'《([^》]+)》', rd['title'])
    cands = [norm(x) for x in inner] or [norm(rd['title'])]
    for d in docs:
        dn = norm(d['name'])
        if len(dn) < 8:
            continue
        for c in cands:
            if len(c) >= 8 and (c in dn or dn in c):
                d['reads'].append(rd['id'])
                hit += 1
                break

body = (
    "        // ===== 政策库（来源：命题方《政策智能体数据库.xlsx》，由脚本结构化，未改一字标题）=====\n"
    "        // 库一 = 全文导入的政策原文 %d 份；库二 = URL 接入的政策与官方解读 %d 条。\n"
    "        // level / tags / type 由标题派生用于匹配；额度、截止日、申报条件原表没有，一律不生成。\n"
    "        const POLICY_DOCS = %s;\n\n"
    "        const POLICY_READS = %s;\n"
) % (
    len(docs), len(reads),
    json.dumps(docs, ensure_ascii=False, indent=0).replace('\n', ''),
    json.dumps(reads, ensure_ascii=False, indent=0).replace('\n', ''),
)

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(body)

print('原文库', len(docs), '解读库', len(reads), '挂接成功', hit)
print('层级分布', {lv: sum(1 for d in docs if d['level'] == lv) for lv in set(d['level'] for d in docs)})
tagcount = {}
for d in docs + reads:
    for t in d['tags']:
        tagcount[t] = tagcount.get(t, 0) + 1
print('标签分布', dict(sorted(tagcount.items(), key=lambda x: -x[1])))
print('可申报事项', sum(1 for r in reads if r['actionable']))
print('输出', OUT, os.path.getsize(OUT), 'bytes')
